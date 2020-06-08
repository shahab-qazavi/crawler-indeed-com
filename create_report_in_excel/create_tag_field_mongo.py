from openpyxl import load_workbook
from pymongo import MongoClient

con = MongoClient()
db = con['crawling_indeed']
col_tags = db['tags']
col_crawl = db['crawling']


wb = load_workbook(filename='tag_name.xlsx')
sheet = wb.active

col_tags.delete_many({})


def excel_save_db():

    for a in range(1, len(sheet.rows)):
        if sheet.cell(row=a, column=1).value is None:
            col_tags.insert_one({
                'name': sheet.cell(row=a, column=1).value,
                'cont': 0
            })


excel_save_db()

