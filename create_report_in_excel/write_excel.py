from openpyxl import Workbook


skill = Workbook()
skill_sheet = skill.active


def write_excel(file_excel_name,database_name):
    r = 1
    for item in database_name.find({}, {'_id': 0}).sort('cont', -1):
        c = 1
        skill_sheet.cell(row=r, column=c).value = item['name']
        c += 1
        skill_sheet.cell(row=r, column=c).value = item['cont']
        r += 1
    skill.save(f'{file_excel_name}.xlsx')
